# implementing a support vector machine classifier to identify the hand position
import openpyxl as xl
import numpy as np
from sklearn import svm
from sklearn.preprocessing import StandardScaler

from sklearn.model_selection import cross_val_score
from sklearn.linear_model import LogisticRegression
import pickle

# Get all data into one dataset from the disjointed excel files collected with sensory-glove
if 1:
    sources = ['open1', 'open2', 'open3', 'open4', 'closed1', 'closed2', 'closed3', 'closed4', 'start1', 'start2', 'start3']
    target_file = 'data'
    target_wb = xl.Workbook()
    target_ws = target_wb.active
    for source in sources:

        wb = xl.load_workbook(f'data/{source}.xlsx')
        ws = wb.active
        num_rows = ws.max_row
        #target_ws.append(content)
        print(ws.max_row)
        for row in ws.iter_rows(min_row=1,
                                   max_row=num_rows,
                                   values_only=True):
            target_ws.append(row)

    target_wb.save(f'data/{target_file}.xlsx')

if 1:
    wb = xl.load_workbook(f'data/data.xlsx')
    ws = wb.active
    num_rows = ws.max_row
    counter = {'open': 0, 'closed': 0, 'start': 0}
    for row in ws.iter_rows(min_row=1,
                               max_row=num_rows,
                               values_only=True):
        if row[-1] == 'open':
            counter['open'] += 1
        elif row[-1] == 'closed':
            counter['closed'] += 1
        elif row[-1] == 'start':
            counter['start'] += 1
        else:
            print('Error')
    print(counter) # {'open': 4244, 'closed': 4738, 'start': 5510}

# Shuffle the data, split into train and UNSEEN test data
if 1:
    wb = xl.load_workbook(f'data/data.xlsx')
    ws = wb.active
    num_rows = ws.max_row
    content = []
    for row in ws.iter_rows(min_row=1,
                               max_row=num_rows,
                               values_only=True):
    content.append(row)
    content = np.array(content)
    np.random.shuffle(content)
    content_train = content[:int(0.8*len(content))]
    content_test = content[int(0.8*len(content)):]
    print(f'Length of train data: {len(content_train)}, length of test data: {len(content_test)}')

    shuffled_train_wb = xl.Workbook()
    shuffled_train_ws = shuffled_train_wb.active
    for row in content_train:
        shuffled_train_ws.append(list(row))
    shuffled_train_wb.save(f'data/shuffled_data_train.xlsx')
    for row in content_test:
        shuffled_train_ws.append(list(row))
    shuffled_train_wb.save(f'data/shuffled_data_test.xlsx')




if __name__ == '__main__':
    svm = svm.SVC(kernel='rbf', C=1)
    log = LogisticRegression()
    # Choose the model
    model = svm
    pipeline = (StandardScaler(), model)

    wb = xl.load_workbook(f'data/shuffled_data_train.xlsx')
    ws = wb.active
    data = []
    labels = []
    for rows in ws.iter_rows(min_row=1,
                               max_row=ws.max_row,
                               values_only=True):
        data.append(rows[:6]) # only 6 sensors
        labels.append(rows[-1])

    # cross validation
    scores = cross_val_score(model, data, labels, cv=4)
    print(f'Accuracy: {scores.mean():.2f} (+/- {scores.std() * 2:.2f})')
    pickle.dump(model, open('svm_model.sav', 'wb'))
